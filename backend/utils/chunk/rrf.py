# -*- coding: utf-8 -*-
import os
import warnings
warnings.filterwarnings('ignore')

import pandas as pd
from openpyxl import load_workbook
import fitz  # PyMuPDF
from PIL import Image
import io
from tqdm import tqdm
import logging
from pathlib import Path
import hashlib
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple, Optional
import numpy as np
from scipy import stats

# LangChain 相关导入
from langchain_community.document_loaders import DirectoryLoader, PyMuPDFLoader, TextLoader, Docx2txtLoader
from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma
from langchain_openai import ChatOpenAI
from langchain.chains.retrieval_qa.base import RetrievalQA
from langchain.prompts import PromptTemplate
from langchain.schema import Document
from langchain_core.retrievers import BaseRetriever
from pydantic import BaseModel


# Embeddings（优先新的包）
try:
    from langchain_huggingface import HuggingFaceEmbeddings
    print("✅ 使用新的 langchain-huggingface 包")
except ImportError:
    from langchain_community.embeddings import HuggingFaceEmbeddings
    print("⚠️ 使用旧的 langchain_community.embeddings，建议升级: pip install langchain-huggingface")

# 词汇相似度 / 去重 依赖
try:
    import jieba
    from rank_bm25 import BM25Okapi
    from rapidfuzz import fuzz

    # 加载涉藏舆情领域自定义词典，提升 BM25 分词精度
    _user_dict = Path(__file__).resolve().parent.parent.parent / "jieba_user_dict.txt"
    if _user_dict.exists():
        jieba.load_userdict(str(_user_dict))
        # suggest_freq 处理高词频冲突，确保领域词不被内置词典覆盖
        for _term in ("藏文化",):
            jieba.suggest_freq(_term, tune=True)
except Exception as _e:
    print("⚠️ 缺少依赖：请先安装 `rank-bm25 jieba rapidfuzz`：")
    print("   pip install rank-bm25 jieba rapidfuzz")
    raise

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 创建模型目录
MODELS_DIR = "./models"
os.makedirs(MODELS_DIR, exist_ok=True)

# 模型配置
MODEL_CONFIG = {
    "embedding": {
        "name": "BAAI/bge-large-zh-v1.5",
        "local_path": os.path.join(MODELS_DIR, "bge-large-zh-v1.5")
    },
    "reranker": {
        "name": "BAAI/bge-reranker-large",
        "local_path": os.path.join(MODELS_DIR, "bge-reranker-large")
    }
}

# --------------------- 设备/模型下载 ---------------------
def check_cuda_availability():
    """检查CUDA是否可用"""
    try:
        import torch
        if torch.cuda.is_available():
            print(f"✅ CUDA 可用，GPU: {torch.cuda.get_device_name(0)}")
            print(f"   CUDA 版本: {torch.version.cuda}")
            return True
        else:
            print("❌ CUDA 不可用，将使用 CPU")
            return False
    except Exception as e:
        print(f"❌ 检查 CUDA 时出错: {e}")
        return False

def get_compatible_device():
    """获取兼容的设备"""
    try:
        import torch
        if torch.cuda.is_available():
            capability = torch.cuda.get_device_capability(0)
            print(f"✅ GPU 检测: {torch.cuda.get_device_name(0)}")
            print(f"✅ CUDA 计算能力: {capability}")
            print(f"✅ CUDA 版本: {torch.version.cuda}")
            return 'cuda'
        else:
            print("❌ CUDA 不可用，使用 CPU")
            return 'cpu'
    except Exception as e:
        print(f"❌ 设备检测失败: {e}, 使用 CPU")
        return 'cpu'

def setup_models():
    """设置和管理本地模型"""
    print("🔧 正在设置模型...")
    for model_type, config in MODEL_CONFIG.items():
        os.makedirs(config["local_path"], exist_ok=True)
        print(f"📁 模型目录: {config['local_path']}")

    embedding_model_path = os.path.join(MODEL_CONFIG["embedding"]["local_path"], "pytorch_model.bin")
    reranker_model_path = os.path.join(MODEL_CONFIG["reranker"]["local_path"], "pytorch_model.bin")

    if os.path.exists(embedding_model_path) and os.path.exists(reranker_model_path):
        print("✅ 模型文件已存在，跳过下载")
        return True

    print("⚠️  模型文件不存在，首次运行需要下载")
    print("📥 请确保网络连接正常，下载可能需要一些时间...")

    try:
        from huggingface_hub import snapshot_download
        os.environ["HF_HUB_DISABLE_PROGRESS_BARS"] = "1"

        if not os.path.exists(embedding_model_path):
            print(f"📥 正在下载嵌入模型: {MODEL_CONFIG['embedding']['name']}")
            snapshot_download(
                repo_id=MODEL_CONFIG["embedding"]["name"],
                local_dir=MODEL_CONFIG["embedding"]["local_path"],
                local_dir_use_symlinks=False,
                resume_download=True
            )
            print("✅ 嵌入模型下载完成")

        if not os.path.exists(reranker_model_path):
            print(f"📥 正在下载重排序模型: {MODEL_CONFIG['reranker']['name']}")
            snapshot_download(
                repo_id=MODEL_CONFIG["reranker"]["name"],
                local_dir=MODEL_CONFIG["reranker"]["local_path"],
                local_dir_use_symlinks=False,
                resume_download=True
            )
            print("✅ 重排序模型下载完成")
        return True

    except ImportError:
        print("❌ 请先安装 huggingface_hub: pip install huggingface-hub")
        return False
    except Exception as e:
        print(f"❌ 模型下载失败: {e}")
        return False

# --------------------- 加载器 ---------------------
class EnhancedExcelLoader:
    def __init__(self):
        pass

    def load_excel_file(self, file_path):
        """增强版Excel文件加载，更好地处理政务表格数据"""
        documents = []
        try:
            # 方法1: 使用pandas读取数据
            try:
                excel_file = pd.ExcelFile(file_path)
                for sheet_name in excel_file.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    text_content = f"表格: {sheet_name}\n"
                    text_content += df.to_string(index=False)
                    doc = Document(
                        page_content=text_content,
                        metadata={
                            "source": file_path,
                            "sheet": sheet_name,
                            "type": "excel_table",
                            "rows": len(df),
                            "columns": len(df.columns) if len(df) > 0 else 0
                        }
                    )
                    documents.append(doc)
            except Exception as e:
                logger.warning(f"Pandas读取Excel失败 {file_path}: {e}")
                # 方法2: 使用openpyxl直接读取
                wb = load_workbook(file_path)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text_content = f"表格: {sheet_name}\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                        text_content += row_text + "\n"
                    doc = Document(
                        page_content=text_content,
                        metadata={
                            "source": file_path,
                            "sheet": sheet_name,
                            "type": "excel_raw",
                            "method": "openpyxl"
                        }
                    )
                    documents.append(doc)
        except Exception as e:
            logger.error(f"Excel文件处理完全失败 {file_path}: {e}")

        return documents

    def load(self, file_path):
        """统一的加载接口"""
        return self.load_excel_file(file_path)

class EnhancedPDFLoader:
    def __init__(self):
        self.standard_loader = PyMuPDFLoader

    def load(self, file_path):
        """仅使用标准PDF解析，禁用OCR功能"""
        try:
            standard_loader = self.standard_loader(file_path)
            documents = standard_loader.load()
            if documents and any(doc.page_content.strip() for doc in documents):
                return documents
            else:
                logger.warning(f"PDF文件无法解析（可能是扫描版）: {file_path}")
                logger.info("💡 OCR功能已禁用，如需处理扫描版PDF请后续安装兼容的PaddlePaddle版本")
                return []
        except Exception as e:
            logger.warning(f"PDF解析失败 {file_path}: {e}")
            return []

def load_documents(data_dir="./data"):
    documents = []
    excel_extensions = ['.xlsx', '.xls']
    for ext in excel_extensions:
        excel_files = [f for f in os.listdir(data_dir) if f.lower().endswith(ext)]
        for excel_file in tqdm(excel_files, desc=f"加载{ext.upper()}文件"):
            excel_path = os.path.join(data_dir, excel_file)
            try:
                loader = EnhancedExcelLoader()
                excel_docs = loader.load(excel_path)
                if excel_docs:
                    documents.extend(excel_docs)
                    logger.info(f"成功加载Excel文件: {excel_file}, 生成 {len(excel_docs)} 个文档")
                else:
                    logger.warning(f"Excel文件没有内容: {excel_file}")
            except Exception as e:
                logger.error(f"加载Excel文件失败 {excel_file}: {e}")

    pdf_files = [f for f in os.listdir(data_dir) if f.lower().endswith('.pdf')]
    for pdf_file in tqdm(pdf_files, desc="加载PDF文件"):
        pdf_path = os.path.join(data_dir, pdf_file)
        try:
            loader = EnhancedPDFLoader()
            pdf_docs = loader.load(pdf_path)
            if pdf_docs:
                documents.extend(pdf_docs)
                logger.info(f"成功加载PDF文件: {pdf_file}, 生成 {len(pdf_docs)} 个文档")
            else:
                logger.warning(f"PDF文件无法解析或为空: {pdf_file}")
        except Exception as e:
            logger.error(f"加载PDF文件失败 {pdf_file}: {e}")

    text_loaders = {
        '.txt': TextLoader,
        '.docx': Docx2txtLoader,
    }

    for ext, loader_cls in text_loaders.items():
        try:
            loader = DirectoryLoader(data_dir, glob=f"**/*{ext}", loader_cls=loader_cls)
            docs = loader.load()
            documents.extend(docs)
            logger.info(f"成功加载 {ext.upper()} 文件: {len(docs)} 个文档")
        except Exception as e:
            logger.warning(f"加载 {ext} 文件时出错: {e}")

    print(f"已成功加载 {len(documents)} 篇文档")
    return documents

def split_documents(documents):
    excel_docs = [doc for doc in documents if doc.metadata.get('type', '').startswith('excel')]
    other_docs = [doc for doc in documents if not doc.metadata.get('type', '').startswith('excel')]

    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=800,
        chunk_overlap=80,
        length_function=len,
        separators=["\n\n", "\n", "。", "！", "？", "；", "，", "、", " "]
    )

    all_texts = []
    if other_docs:
        texts = text_splitter.split_documents(other_docs)
        all_texts.extend(texts)
        print(f"已分割 {len(other_docs)} 篇非Excel文档成 {len(texts)} 个文本块")

    if excel_docs:
        for excel_doc in excel_docs:
            if len(excel_doc.page_content) > 1500:
                chunks = text_splitter.split_documents([excel_doc])
                all_texts.extend(chunks)
            else:
                all_texts.append(excel_doc)
        print(f"已处理 {len(excel_docs)} 个Excel表格文档")

    print(f"总共生成 {len(all_texts)} 个文本块")
    return all_texts

# --------------------- 向量库 ---------------------
def create_embeddings():
    """创建使用本地模型的嵌入实例"""
    device = get_compatible_device()
    return HuggingFaceEmbeddings(
        model_name=MODEL_CONFIG["embedding"]["local_path"],
        model_kwargs={'device': device},
        encode_kwargs={'normalize_embeddings': True, 'batch_size': 64}
    )

def create_vectorstore(texts):
    embeddings = create_embeddings()
    vectorstore = Chroma.from_documents(
        documents=texts,
        embedding=embeddings,
        persist_directory="./chroma_db"
    )
    vectorstore.persist()
    print("✅ 向量数据库已创建并保存到 './chroma_db'")
    return vectorstore

def validate_vectorstore(vectorstore):
    """验证向量数据库是否正常可用"""
    try:
        test_results = vectorstore.similarity_search("测试", k=1)
        if test_results:
            print("✅ 向量数据库验证成功")
            return True
        else:
            print("❌ 向量数据库为空或异常")
            return False
    except Exception as e:
        print(f"❌ 向量数据库验证失败: {e}")
        return False

# --------------------- LLM ---------------------
def create_deepseek_llm():
    """创建DeepSeek LLM实例（兼容OpenAI API）"""
    api_key = os.environ.get("DEEPSEEK_API_KEY")
    if not api_key:
        print("❌ 请设置 DEEPSEEK_API_KEY 环境变量")
        print("💡 设置方法: export DEEPSEEK_API_KEY='你的-api-key'")
        return None

    return ChatOpenAI(
        model="deepseek-chat",
        base_url="https://api.deepseek.com/v1",
        api_key=api_key,
        temperature=0,
        max_tokens=2000,
        timeout=30
    )

# --------------------- 增强检索算法 ---------------------
def _stable_uid(doc: Document) -> str:
    meta = doc.metadata or {}
    source = str(meta.get("source", ""))
    sheet = str(meta.get("sheet", ""))
    page = str(meta.get("page", ""))
    head = (doc.page_content or "")[:256]
    return hashlib.md5(f"{source}||{sheet}||{page}||{head}".encode("utf-8")).hexdigest()

class BM25Indexer:
    """简易 BM25 索引器（内存版，可持久化到 ./bm25_store）"""
    def __init__(self, store_dir="./bm25_store"):
        self.store_dir = store_dir
        self._bm25: BM25Okapi = None
        self._doc_ids: List[str] = []
        self._doc_map: Dict[str, Dict[str, Any]] = {}
        self._corpus_tokens: List[List[str]] = []

    def _build_text_for_bm25(self, doc: Document) -> str:
        meta = doc.metadata or {}
        parts = []
        if meta.get("sheet"): parts.append(str(meta["sheet"]))
        if meta.get("source"): parts.append(str(meta["source"]))
        parts.append(doc.page_content or "")
        return "\n".join(parts)

    def fit(self, documents: List[Document]):
        self._doc_ids, self._doc_map, self._corpus_tokens = [], {}, []
        for doc in documents:
            txt = self._build_text_for_bm25(doc)
            toks = jieba.lcut(txt)
            uid = _stable_uid(doc)
            self._doc_ids.append(uid)
            self._doc_map[uid] = {"text": doc.page_content or "", "meta": doc.metadata or {}}
            self._corpus_tokens.append(toks)
        if not self._corpus_tokens:
            raise RuntimeError("BM25Indexer.fit: empty corpus")
        self._bm25 = BM25Okapi(self._corpus_tokens)

    def save(self):
        os.makedirs(self.store_dir, exist_ok=True)
        import pickle
        with open(os.path.join(self.store_dir, "bm25.pkl"), "wb") as f:
            import pickle as p; p.dump(self._bm25, f)
        with open(os.path.join(self.store_dir, "doc_ids.pkl"), "wb") as f:
            import pickle as p; p.dump(self._doc_ids, f)
        with open(os.path.join(self.store_dir, "doc_map.pkl"), "wb") as f:
            import pickle as p; p.dump(self._doc_map, f)

    def load(self):
        import pickle
        with open(os.path.join(self.store_dir, "bm25.pkl"), "rb") as f:
            self._bm25 = pickle.load(f)
        with open(os.path.join(self.store_dir, "doc_ids.pkl"), "rb") as f:
            self._doc_ids = pickle.load(f)
        with open(os.path.join(self.store_dir, "doc_map.pkl"), "rb") as f:
            self._doc_map = pickle.load(f)

    def try_load(self) -> bool:
        try:
            self.load()
            return True
        except Exception:
            return False

    def is_ready(self) -> bool:
        return self._bm25 is not None and len(self._doc_ids) > 0

    def search(self, query: str, k: int = 50) -> List[Tuple[str, float]]:
        if not self.is_ready():
            raise RuntimeError("BM25Indexer not ready. Call fit()/load() first.")
        q_tokens = jieba.lcut(query or "")
        scores = self._bm25.get_scores(q_tokens)
        import numpy as np
        scores = np.array(scores, dtype=float)
        idx = scores.argsort()[::-1][:k]
        return [(self._doc_ids[int(i)], float(scores[int(i)])) for i in idx]

    def id_to_document(self, doc_id: str) -> Document:
        item = self._doc_map[doc_id]
        return Document(page_content=item["text"], metadata=item["meta"])

@dataclass
class Cand:
    doc: Document
    cos: float = 0.0
    bm25: float = 0.0
    rrf_score: float = 0.0
    s_ce: float = 0.0
    s_ce_norm: float = 0.0
    mmr_score: float = 0.0

class RRFMerger:
    """RRF融合器"""
    
    @staticmethod
    def reciprocal_rank_fusion(vec_results: List[Tuple[Document, float]], 
                              bm25_results: List[Tuple[str, float]], 
                              bm25_indexer: BM25Indexer,
                              k: int = 10) -> List[Tuple[Document, float]]:  # 将k从60改为10
        """
        RRF融合算法
        """
        doc_rank_map = {}
        
        # 处理向量检索结果
        for rank, (doc, score) in enumerate(vec_results):
            uid = _stable_uid(doc)
            if uid not in doc_rank_map:
                doc_rank_map[uid] = {'doc': doc, 'ranks': []}
            # 确保排名从1开始
            doc_rank_map[uid]['ranks'].append(rank + 1)
        
        # 处理BM25检索结果
        for rank, (doc_id, score) in enumerate(bm25_results):
            try:
                doc = bm25_indexer.id_to_document(doc_id)
                uid = _stable_uid(doc)
                if uid not in doc_rank_map:
                    doc_rank_map[uid] = {'doc': doc, 'ranks': []}
                doc_rank_map[uid]['ranks'].append(rank + 1)
            except Exception as e:
                continue
        
        # 计算RRF分数
        rrf_scores = []
        for uid, info in doc_rank_map.items():
            rrf_score = 0.0
            for rank in info['ranks']:
                rrf_score += 1.0 / (k + rank)  # 使用k=10而不是60
            rrf_scores.append((info['doc'], rrf_score))
        
        # 按RRF分数降序排序
        rrf_scores.sort(key=lambda x: x[1], reverse=True)
        return rrf_scores

class AdaptiveThresholdFilter:
    """自适应阈值筛选器"""
    
    @staticmethod
    def find_optimal_threshold(scores: List[float], 
                             min_docs: int = 3,
                             quantile: float = 0.5,
                             min_gap_ratio: float = 0.1) -> Tuple[float, List[int]]:
        """
        自适应阈值：分位数 + Gap截断 + 缺证兜底
        """
        if not scores:
            return 0.0, []
            
        scores_array = np.array(scores)
        
        # 1. 分位数基准
        quantile_threshold = np.quantile(scores_array, quantile)
        
        # 2. Gap截断分析
        gaps = []
        for i in range(1, len(scores_array)):
            gap = scores_array[i-1] - scores_array[i]
            gap_ratio = gap / (scores_array[0] + 1e-8)  # 相对第一个分数的比例
            gaps.append((i, gap, gap_ratio))
        
        # 寻找显著Gap
        gap_threshold_idx = len(scores_array) - 1  # 默认取最后一个
        for idx, gap, gap_ratio in gaps:
            if gap_ratio > min_gap_ratio and idx >= min_docs:
                gap_threshold_idx = idx
                break
        
        gap_threshold = scores_array[gap_threshold_idx] if gap_threshold_idx < len(scores_array) else scores_array[-1]
        
        # 3. 最终阈值：取分位数和Gap阈值的最大值，确保至少min_docs个文档
        final_threshold = max(quantile_threshold, gap_threshold)
        
        # 确保至少保留min_docs个文档
        keep_indices = []
        for i, score in enumerate(scores_array):
            if score >= final_threshold or len(keep_indices) < min_docs:
                keep_indices.append(i)
        
        return final_threshold, keep_indices

class MMRDiversityReranker:
    """修复版MMR多样性重排器"""
    
    def __init__(self, embeddings, lambda_param: float = 0.7):
        self.embeddings = embeddings
        self.lambda_param = lambda_param
    
    def compute_mmr_scores(self, query: str, candidates: List[Cand], top_k: int) -> List[Cand]:
        if len(candidates) <= top_k:
            for cand in candidates:
                cand.mmr_score = cand.s_ce_norm
            return candidates
        
        # 嵌入获取
        try:
            query_embedding = np.array(self.embeddings.embed_query(query))
        except Exception as e:
            # 回退到CE分数排序
            sorted_cands = sorted(candidates, key=lambda x: x.s_ce_norm, reverse=True)[:top_k]
            for cand in sorted_cands:
                cand.mmr_score = cand.s_ce_norm
            return sorted_cands
        
        # 文档嵌入获取
        doc_embeddings = []
        valid_candidates = []
        
        for cand in candidates:
            try:
                content = cand.doc.page_content or ""
                if len(content.strip()) > 50:  # 确保内容足够长
                    doc_embedding = np.array(self.embeddings.embed_documents([content])[0])
                    doc_embeddings.append(doc_embedding)
                    valid_candidates.append(cand)
            except Exception as e:
                continue
        
        if len(valid_candidates) <= top_k:
            for cand in valid_candidates:
                cand.mmr_score = cand.s_ce_norm
            return valid_candidates
        
        # MMR算法
        selected = []
        remaining = list(range(len(valid_candidates)))
        
        # 按CE分数选择第一个
        first_idx = max(remaining, key=lambda i: valid_candidates[i].s_ce_norm)
        selected.append(first_idx)
        remaining.remove(first_idx)
        valid_candidates[first_idx].mmr_score = valid_candidates[first_idx].s_ce_norm
        
        while len(selected) < min(top_k, len(valid_candidates)) and remaining:
            best_score = -float('inf')
            best_idx = -1
            
            for idx in remaining:
                relevance = valid_candidates[idx].s_ce_norm
                
                # 相似度计算
                max_similarity = 0.0
                for sel_idx in selected:
                    try:
                        vec1 = doc_embeddings[idx]
                        vec2 = doc_embeddings[sel_idx]
                        # 确保向量归一化
                        vec1_norm = vec1 / (np.linalg.norm(vec1) + 1e-8)
                        vec2_norm = vec2 / (np.linalg.norm(vec2) + 1e-8)
                        similarity = np.dot(vec1_norm, vec2_norm)
                        max_similarity = max(max_similarity, similarity)
                    except:
                        continue
                
                # MMR分数计算
                mmr_score = np.clip(self.lambda_param * relevance - (1 - self.lambda_param) * max_similarity, 0.0, 1.0)
                
                if mmr_score > best_score:
                    best_score = mmr_score
                    best_idx = idx
            
            if best_idx != -1:
                selected.append(best_idx)
                remaining.remove(best_idx)
                valid_candidates[best_idx].mmr_score = best_score
            else:
                break
        
        result = [valid_candidates[i] for i in selected]
        
        # 确保所有候选都有合理的MMR分数
        for cand in result:
            if not hasattr(cand, 'mmr_score') or cand.mmr_score < 0:
                cand.mmr_score = cand.s_ce_norm
        
        return result

class CrossEncoderWrapper:
    """使用 sentence-transformers.CrossEncoder（bge-reranker-large）"""
    def __init__(self, model_path: str, batch_size: int = 16):
        from sentence_transformers import CrossEncoder
        import torch
        self.batch_size = batch_size
        device = "cuda" if torch.cuda.is_available() else "cpu"
        
        # 兼容不同版本的初始化方式
        try:
            self.model = CrossEncoder(model_path, device=device)
        except Exception as e:
            # 如果直接传递失败，尝试其他方式
            try:
                self.model = CrossEncoder(model_name_or_path=model_path, device=device)
            except:
                self.model = CrossEncoder(model_name=model_path, device=device)

    def score(self, query: str, passages: List[str]) -> List[float]:
        if not passages:
            return []
        pairs = [(query, p) for p in passages]
        scores = self.model.predict(pairs, batch_size=self.batch_size)
        return [float(s) for s in scores]


class EnhancedHybridRetriever:
    """
    增强版混合检索器：RRF融合 + 自适应阈值 + MMR多样性重排
    """
    def __init__(
        self,
        vectorstore,
        bm25_indexer: BM25Indexer,
        cross_encoder: CrossEncoderWrapper,
        embeddings,
        *,
        rrf_k: int = 60,
        m_for_rerank: int = 80,
        top_k: int = 6,
        min_docs: int = 3,
        quantile_threshold: float = 0.5,
        min_gap_ratio: float = 0.1,
        mmr_lambda: float = 0.7,
        dup_text_threshold: int = 90,
        per_source_max: int = 1,
        meta_exact_boost: float = 0.03,
        search_kwargs: dict = None
    ):
        self.vs = vectorstore
        self.bm25 = bm25_indexer
        self.ce = cross_encoder
        self.embeddings = embeddings
        
        # RRF参数
        self.rrf_k = rrf_k
        
        # 重排序参数
        self.m_for_rerank = m_for_rerank
        self.top_k = top_k
        self.min_docs = min_docs
        
        # 自适应阈值参数
        self.quantile_threshold = quantile_threshold
        self.min_gap_ratio = min_gap_ratio
        
        # MMR参数
        self.mmr_lambda = mmr_lambda
        self.mmr_reranker = MMRDiversityReranker(embeddings, mmr_lambda)
        
        # 去重参数
        self.dup_text_threshold = dup_text_threshold
        self.per_source_max = per_source_max
        self.meta_exact_boost = meta_exact_boost
        self.search_kwargs = search_kwargs or {}

    # LangChain 适配
    def get_relevant_documents(self, query: str) -> List[Document]:
        cands = self._retrieve(query)
        out = []
        for c in cands:
            meta = dict(c.doc.metadata or {})
            meta.update({
                "_cos": round(c.cos, 6),
                "_bm25": round(c.bm25, 6),
                "_rrf_score": round(c.rrf_score, 6),
                "_s_ce": round(c.s_ce, 6),
                "_s_ce_norm": round(c.s_ce_norm, 6),
                "_mmr_score": round(c.mmr_score, 6),
            })
            out.append(Document(page_content=c.doc.page_content, metadata=meta))
        return out

    # 主流程
    def _retrieve(self, query: str) -> List[Cand]:
        # 1) 语义检索
        vec_hits = []
        try:
            vec_hits = self.vs.similarity_search_with_score(query, k=self.m_for_rerank, **self.search_kwargs)
        except Exception:
            docs = self.vs.similarity_search(query, k=self.m_for_rerank, **self.search_kwargs)
            vec_hits = [(d, 0.0) for d in docs]

        # 2) BM25检索
        bm25_hits = self.bm25.search(query, k=self.m_for_rerank)

        # 3) RRF融合
        rrf_results = RRFMerger.reciprocal_rank_fusion(vec_hits, bm25_hits, self.bm25, self.rrf_k)
        # 添加调试信息
        if rrf_results:
            max_rrf = max(score for _, score in rrf_results)
            min_rrf = min(score for _, score in rrf_results)
            logger.info(f"RRF分数范围: {min_rrf:.4f} - {max_rrf:.4f}")
        
        # 4) 构建候选文档
        cands = []
        for doc, rrf_score in rrf_results[:self.m_for_rerank]:
            # 找到原始分数
            cos_score = 0.0
            bm25_score = 0.0
            
            for vec_doc, score in vec_hits:
                if _stable_uid(vec_doc) == _stable_uid(doc):
                    cos_score = 1.0 - float(score) if score <= 1.0 else float(score)
                    break
            
            for doc_id, score in bm25_hits:
                if doc_id == _stable_uid(doc):
                    bm25_score = float(score)
                    break
            
            cand = Cand(
                doc=doc,
                cos=cos_score,
                bm25=bm25_score,
                rrf_score=rrf_score
            )
            cands.append(cand)

        if not cands:
            return []

        # 5) 交叉编码器重排序
        ce_scores = self.ce.score(query, [c.doc.page_content or "" for c in cands])
        for cand, score in zip(cands, ce_scores):
            cand.s_ce = float(score)

        # 6) CE分数归一化
        ce_scores_list = [c.s_ce for c in cands]
        min_ce = min(ce_scores_list)
        max_ce = max(ce_scores_list)
        
        if max_ce > min_ce:
            for cand in cands:
                cand.s_ce_norm = (cand.s_ce - min_ce) / (max_ce - min_ce)
        else:
            for i, cand in enumerate(cands):
                cand.s_ce_norm = 0.5 + (i - len(cands)/2) * 0.01

        # 7) 自适应阈值筛选
        threshold, keep_indices = AdaptiveThresholdFilter.find_optimal_threshold(
            [c.s_ce_norm for c in cands], 
            self.min_docs, 
            self.quantile_threshold,
            self.min_gap_ratio
        )
        
        filtered_cands = [cands[i] for i in keep_indices]

        # 8) 去重
        deduped_cands = self._dedupe(filtered_cands)

        # 9) MMR多样性重排
        final_cands = self.mmr_reranker.compute_mmr_scores(query, deduped_cands, self.top_k)

        return final_cands[:self.top_k]

    def _dedupe(self, ranked: List[Cand]) -> List[Cand]:
        """去重逻辑"""
        kept: List[Cand] = []
        src_counter: Dict[Tuple[str, str, str], int] = {}

        for cand in ranked:
            meta = cand.doc.metadata or {}
            key = (str(meta.get("source", "")),
                   str(meta.get("sheet", "")),
                   str(meta.get("page", "")))
            # 来源级限制
            cnt = src_counter.get(key, 0)
            if cnt >= self.per_source_max:
                continue

            # 文本近重复
            text = cand.doc.page_content or ""
            dup = False
            for k in kept:
                if fuzz.token_set_ratio(text, k.doc.page_content or "") >= self.dup_text_threshold:
                    dup = True
                    break
            if dup:
                continue

            kept.append(cand)
            src_counter[key] = cnt + 1

            if len(kept) >= max(self.top_k * 2, self.min_docs * 2):
                break

        return kept
    
# --- LangChain 兼容包装器 ---
class EnhancedHybridLCRetriever(BaseRetriever):
    """将增强版 HybridRetriever 适配为 LangChain 的 BaseRetriever。"""

    def __init__(self, hybrid: "EnhancedHybridRetriever", **kwargs):
        super().__init__(**kwargs)
        self._hybrid = hybrid

    def _get_relevant_documents(self, query: str, *, run_manager=None) -> List[Document]:
        return self._hybrid.get_relevant_documents(query)

    async def _aget_relevant_documents(self, query: str, *, run_manager=None) -> List[Document]:
        import anyio
        return await anyio.to_thread.run_sync(self._hybrid.get_relevant_documents, query)

# --------------------- QA 链 ---------------------
def create_qa_chain(retriever):
    llm = create_deepseek_llm()
    if llm is None:
        return None

    prompt_template = """你是一名专业的政务咨询助手，负责解答关于某市政策文件和数据的疑问。

重要要求：
1. 仔细分析提供的上下文，特别是表格数据
2. 如果上下文中有表格，请准确解释表格内容
3. 如果问题涉及数值计算，基于表格数据进行计算
4. 保持回答的准确性和权威性
5. 如果上下文信息不足，请如实告知无法回答

上下文内容：
{context}

问题：{question}

请基于以上信息提供准确解答："""

    PROMPT = PromptTemplate(template=prompt_template, input_variables=["context", "question"])

    qa_chain = RetrievalQA.from_chain_type(
        llm=llm,
        chain_type="stuff",
        retriever=retriever,
        chain_type_kwargs={"prompt": PROMPT},
        return_source_documents=True
    )
    return qa_chain

# --------------------- 交互 ---------------------
def interactive_qa(qa_chain):
    """交互式问答循环"""
    if qa_chain is None:
        print("❌ QA链创建失败，无法进入问答模式")
        return

    print("\n" + "="*60)
    print("增强版政务RAG系统已就绪！")
    print("✅ RRF融合 | 自适应阈值 | MMR多样性重排")
    print("请输入您的问题（输入'quit'退出）")
    print("="*60)

    while True:
        try:
            question = input("\n📋 问题: ").strip()
            if question.lower() == 'quit':
                break
            if not question:
                continue

            print("🔍 正在检索和分析...")
            result = qa_chain.invoke({"query": question})

            print(f"\n✅ 答案: {result['result']}")
            if result.get('source_documents'):
                print(f"\n📚 参考来源 ({len(result['source_documents'])} 篇):")
                for i, doc in enumerate(result['source_documents']):
                    source_info = doc.metadata.get('source', '未知文件')
                    if 'sheet' in doc.metadata and doc.metadata['sheet']:
                        source_info += f" (Sheet: {doc.metadata['sheet']})"
                    elif 'page' in doc.metadata:
                        source_info += f" (第{doc.metadata['page']}页)"
                    
                    # 显示增强的评分信息
                    score_info = []
                    if '_rrf_score' in doc.metadata:
                        score_info.append(f"RRF:{doc.metadata['_rrf_score']:.3f}")
                    if '_s_ce_norm' in doc.metadata:
                        score_info.append(f"CE:{doc.metadata['_s_ce_norm']:.3f}")
                    if '_mmr_score' in doc.metadata:
                        score_info.append(f"MMR:{doc.metadata['_mmr_score']:.3f}")
                    
                    score_str = " | ".join(score_info)
                    print(f"  {i+1}. {os.path.basename(source_info)} [{score_str}]")
        except KeyboardInterrupt:
            print("\n👋 再见！")
            break
        except Exception as e:
            print(f"❌ 处理问题时出错: {e}")

# --------------------- 主程序 ---------------------
def main():
    print("🚀 正在启动增强版政务RAG系统...")

    # 检查CUDA可用性
    cuda_available = check_cuda_availability()
    if not cuda_available:
        print("💡 提示: 安装支持CUDA的PyTorch: pip install torch torchvision torchaudio --index-url https://download.pytorch.org/whl/cu118")

    print("📊 支持格式：PDF, Excel, Word, TXT")
    print("🎯 增强特性：RRF融合 | 自适应阈值 | MMR多样性重排")

    # 设置模型
    if not setup_models():
        print("❌ 模型设置失败，请检查网络连接")
        return

    # 检查/构建向量数据库
    vectorstore = None
    embeddings = create_embeddings()
    
    if os.path.exists("./chroma_db"):
        try:
            vectorstore = Chroma(persist_directory="./chroma_db", embedding_function=embeddings)
            if validate_vectorstore(vectorstore):
                print("✅ 已加载现有向量数据库")
            else:
                vectorstore = None
        except Exception as e:
            print(f"⚠️ 加载现有向量数据库失败: {e}")
            vectorstore = None

    texts = None
    if vectorstore is None:
        print("🔄 正在构建新的向量数据库...")
        documents = load_documents()
        if not documents:
            print("❌ 错误：未加载到任何文档，请检查 data 目录")
            return
        texts = split_documents(documents)
        vectorstore = create_vectorstore(texts)

    # 准备 BM25 索引
    print("🧮 准备 BM25 索引...")
    bm25 = BM25Indexer(store_dir="./bm25_store")
    bm25_ready = False
    try:
        if os.path.exists("./bm25_store/bm25.pkl"):
            bm25.load()
            bm25_ready = True
            print("✅ 已加载现有 BM25 索引")
    except Exception as e:
        print(f"⚠️ 加载 BM25 索引失败：{e}")

    if not bm25_ready:
        if texts is None:
            documents = load_documents()
            if not documents:
                print("❌ 错误：未加载到任何文档，请检查 data 目录")
                return
            texts = split_documents(documents)
        bm25.fit(texts)
        try:
            bm25.save()
            print("✅ 已构建并保存 BM25 索引")
        except Exception as e:
            print(f"⚠️ 保存 BM25 索引失败：{e}")

    # 交叉编码器
    print("⚙️ 创建交叉编码器...")
    ce = CrossEncoderWrapper(MODEL_CONFIG["reranker"]["local_path"])

    # 创建增强版融合检索器
    print("🔄 创建增强版混合检索器...")
    hybrid = EnhancedHybridRetriever(
        vectorstore=vectorstore,
        bm25_indexer=bm25,
        cross_encoder=ce,
        embeddings=embeddings,
        rrf_k=10,
        m_for_rerank=50,
        top_k=6,
        min_docs=3,
        quantile_threshold=0.3,
        min_gap_ratio=0.15,
        mmr_lambda=0.6,
        dup_text_threshold=85,
        per_source_max=1,
        meta_exact_boost=0.05
    )

    # 用 LangChain 兼容包装器包一层
    retriever = EnhancedHybridLCRetriever(hybrid)

    # QA 链
    qa_chain = create_qa_chain(retriever)
    if qa_chain is None:
        print("❌ 创建QA链失败，请检查 DEEPSEEK_API_KEY")
        return

    # 进入交互式问答
    interactive_qa(qa_chain)

if __name__ == "__main__":
    main()